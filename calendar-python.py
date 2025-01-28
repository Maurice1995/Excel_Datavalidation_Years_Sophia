import pandas as pd
from datetime import datetime, timedelta, date
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, Border, Side, PatternFill
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter

def get_easter_date(year):
    """Calculate Easter Sunday date for a given year."""
    a = year % 19
    b = year // 100
    c = year % 100
    d = b // 4
    e = b % 4
    f = (b + 8) // 25
    g = (b - f + 1) // 3
    h = (19 * a + b - d - g + 15) % 30
    i = c // 4
    k = c % 4
    l = (32 + 2 * e + 2 * i - h - k) % 7
    m = (a + 11 * h + 22 * l) // 451
    month = (h + l - 7 * m + 114) // 31
    day = ((h + l - 7 * m + 114) % 31) + 1
    return date(year, month, day)

def get_austrian_holidays(year):
    """Get list of Austrian public holidays for a given year."""
    easter = get_easter_date(year)
    good_friday = easter - timedelta(days=2)
    easter_monday = easter + timedelta(days=1)
    ascension = easter + timedelta(days=39)
    whit_monday = easter + timedelta(days=50)
    corpus_christi = easter + timedelta(days=60)

    holidays = [
        date(year, 1, 1),    # New Year's Day
        date(year, 1, 6),    # Epiphany
        date(year, 5, 1),    # Labor Day
        date(year, 8, 15),   # Assumption Day
        date(year, 10, 26),  # National Day
        date(year, 11, 1),   # All Saints' Day
        date(year, 12, 8),   # Immaculate Conception
        date(year, 12, 25),  # Christmas Day
        date(year, 12, 26),  # St. Stephen's Day
        good_friday,         # Good Friday
        easter_monday,       # Easter Monday
        ascension,          # Ascension Day
        whit_monday,        # Whit Monday
        corpus_christi      # Corpus Christi
    ]
    return holidays

def create_calendar(year):
    wb = Workbook()
    months = {
        1: 'Januar', 2: 'Februar', 3: 'März', 4: 'April',
        5: 'Mai', 6: 'Juni', 7: 'Juli', 8: 'August',
        9: 'September', 10: 'Oktober', 11: 'November', 12: 'Dezember'
    }
    
    weekdays = ['Montag', 'Dienstag', 'Mittwoch', 'Donnerstag', 'Freitag', 'Samstag', 'Sonntag']
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    thick_right_border = Border(left=Side(style='thin'), right=Side(style='thick'), top=Side(style='thin'), bottom=Side(style='thin'))
    bold_font = Font(bold=True)
    grey_fill = PatternFill(start_color='D3D3D3', end_color='D3D3D3', fill_type='solid')
    
    # Get holidays for the year
    holidays = get_austrian_holidays(year)
    
    wb.remove(wb.active)
    
    for month_num in range(1, 13):
        ws = wb.create_sheet(months[month_num])
        
        ws['A1'] = f"{months[month_num]}.{str(year)[-2:]}"
        ws.merge_cells('A1:B1')
        ws['A1'].alignment = Alignment(horizontal='center')
        ws['A1'].font = bold_font
        
        ws['B2'] = "Wochentag"
        ws['B2'].alignment = Alignment(horizontal='center')
        ws['B2'].font = bold_font
        
        ws['C2'] = months[month_num]
        ws['C2'].alignment = Alignment(horizontal='center')
        ws['C2'].font = bold_font
        
        first_day = datetime(year, month_num, 1)
        last_day = datetime(year, month_num + 1, 1) - timedelta(days=1) if month_num < 12 else datetime(year, 12, 31)
        dates = pd.date_range(first_day, last_day)
        
        row = 3
        kw_start_row = None
        current_kw = None
        
        for i, date in enumerate(dates):
            new_kw = date.isocalendar()[1]
            
            if i == 0 or date.weekday() == 0:
                if current_kw != new_kw:
                    if kw_start_row and kw_start_row < row - 1:
                        ws.merge_cells(f'A{kw_start_row}:A{row-1}')
                        ws[f'A{kw_start_row}'].alignment = Alignment(horizontal='center', vertical='center')
                        ws[f'A{kw_start_row}'].font = bold_font
                    
                    ws[f'A{row}'] = f'KW {new_kw}'
                    current_kw = new_kw
                    kw_start_row = row
            
            ws[f'B{row}'] = weekdays[date.weekday()]
            ws[f'C{row}'] = date.day
            ws[f'C{row}'].alignment = Alignment(horizontal='center')
            
            # Create data validation for all rows
            dv = DataValidation(
                type="custom",
                formula1=f'=COUNTIF($D{row}:$AG{row},"u")+COUNTIF($D{row}:$AG{row},"x")<=8',
                error="Für diesen Tag ist das Kontingent bereits aufgebraucht. Kein Eintrag mehr möglich.",
                errorTitle="Kein Eintrag möglich",
                allow_blank=True,
                showErrorMessage=True
            )
            ws.add_data_validation(dv)
            dv.add(f'D{row}:AG{row}')
            
            # Apply grey background for Sundays and holidays
            current_date = date.date()
            if date.weekday() == 6 or current_date in holidays:
                ws[f'A{row}'].fill = grey_fill
                ws[f'B{row}'].fill = grey_fill
                ws[f'C{row}'].fill = grey_fill
                # Grey out all cells in the row
                for col_idx in range(4, 34):  # D to AG
                    col_letter = get_column_letter(col_idx)
                    ws[f'{col_letter}{row}'].fill = grey_fill
            
            ws[f'B{row}'].font = bold_font
            ws[f'C{row}'].font = bold_font
            
            ws[f'A{row}'].border = thin_border
            ws[f'B{row}'].border = thin_border
            ws[f'C{row}'].border = thick_right_border
            
            # Add borders for all cells
            for col_idx in range(4, 34):  # D to AG
                col_letter = get_column_letter(col_idx)
                ws[f'{col_letter}{row}'].border = thin_border
            
            row += 1
        
        if kw_start_row and kw_start_row < row - 1:
            ws.merge_cells(f'A{kw_start_row}:A{row-1}')
            ws[f'A{kw_start_row}'].alignment = Alignment(horizontal='center', vertical='center')
            ws[f'A{kw_start_row}'].font = bold_font
        
        ws['A1'].border = thin_border
        ws['B1'].border = thin_border
        ws['A2'].border = thin_border
        ws['B2'].border = thin_border
        ws['C2'].border = thick_right_border
        
        ws.column_dimensions['A'].width = 8
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 8

    wb.save(f'Kalender_{year}.xlsx')

if __name__ == '__main__':
    year = int(input("Bitte geben Sie das Jahr ein: "))
    create_calendar(year)